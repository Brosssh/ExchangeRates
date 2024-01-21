import datetime
from typing import List

import openpyxl
from openpyxl import load_workbook
import openpyxl.worksheet.worksheet
import openpyxl.utils.exceptions
from openpyxl.cell import Cell

from Classes.consts import MandatoryColumns
from Classes.daily_rate import DailyRates
from Classes.exceptions import MissingColumn
from Managers.utility import try_parse_float


class ExcelManager:
    wb: openpyxl.workbook.workbook.Workbook
    ws :openpyxl.worksheet.worksheet.Worksheet
    file_name :str
    content_start :int
    col_names = {}

    def _build_col_names(self):
        col_names = {}
        current = 0
        for col in self.ws.iter_cols(1, self.ws.max_column):
            col_names[col[0].value] = current
            current += 1
        return col_names

    def _validate_columns(self):
        for mandatory_column in MandatoryColumns.get_names():
            if mandatory_column not in self.col_names:
                raise MissingColumn(mandatory_column)

    def __init__(self, file_path :str) -> None:
        self.wb = load_workbook(filename=file_path)
        self.ws = self.wb.active
        self.file_name = file_path.split("\\")[-1]
        self.max_row = self.ws.max_row
        self.col_names = self._build_col_names()
        self._validate_columns()

    def get_column_by_name(self, name) -> List[Cell]:
        ix = self.col_names[name]
        return self.get_column_by_index(ix)

    def get_column_by_index(self, index) -> List[Cell]:
        return list(self.ws.columns)[index]

    def get_last_date(self) -> datetime.datetime:
        col = self.get_column_by_name(MandatoryColumns.Date.value)
        value_list :List[datetime.datetime] = [x.value for x in col[1:] if x.value is not None and type(x.value) == datetime.datetime]
        if len(value_list)>0:
            return max(value_list) + datetime.timedelta(days=1)
        else:
            return datetime.datetime(2022, 1, 1)

    def append(self, current_date :datetime.datetime, api_response :DailyRates):
        if len(api_response.rates) == 0:
            print(f"INFO: No rates found for {current_date:%Y-%m-%d}.")
            return

        print(f"INFO: Inserting rates for {current_date:%Y-%m-%d}.")
        for r in api_response.rates:
            self.ws.append([current_date, r.isoCode, "EUR", try_parse_float(r.avgRate)])
        self._save_file()

    def _set_new_cells_format(self, column_name :MandatoryColumns, format :str):
        new_max_row: int = self.ws.max_row
        col = self.get_column_by_name(column_name.value)
        for row in range(self.max_row, new_max_row):
            col[row].number_format = format

    def _save_file(self):
        self._set_new_cells_format(MandatoryColumns.Date, 'yyyy-mm-dd;@')
        self._set_new_cells_format(MandatoryColumns.Rate, '0.00000')
        self.wb.save(self.file_name)
        self.max_row = self.ws.max_row

def is_valid_excel_file(file_path :str) -> bool:
    try:
        load_workbook(filename=file_path)
        return True
    except openpyxl.utils.exceptions.InvalidFileException:
        return False

